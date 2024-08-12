from aiogram.types import Message
from aiogram.filters import CommandStart, Command, Filter
from aiogram import F
from loader import bot, router_admin, db_users
from keyboards.defoult.keys import *
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import openpyxl
from aiogram.types.input_file import FSInputFile
from aiogram.enums import ChatAction
import os

workbook = openpyxl.Workbook()
sheet = workbook.active

class Admins(StatesGroup):
    user_id = State()
    name = State()
    surname = State()
    contact = State()
    nikname = State()

class Admin(Filter):
    def __init__(self, my_id: int):
        self.my_id = my_id

    async def __call__(self, msg: Message):
        return msg.from_user.id == self.my_id
    
ADMIN =  "ADMIN_ID"

@router_admin.message(CommandStart(), Admin(ADMIN))
async def start(msg: Message):
    await msg.answer("Assalomu aleykum Admin!\nFoydalanuvchilar haqida bilishi uchun /user buyrug'ini yuboring!")

#-------------------Sotuvchilar Table----------------------

@router_admin.message(Command('user'), Admin(ADMIN))
async def sotuvchilar(msg: Message):
    await msg.bot.send_chat_action(chat_id=msg.chat.id, action=ChatAction.UPLOAD_DOCUMENT)
    
    # Yangi varaq yaratish
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Eski ma'lumotlarni tozalash
    sheet.delete_rows(1, sheet.max_row)

    user_info = db_users.select_users()
    if user_info:
        # Ustun sarlavhalarini qo'shish
        headers = ['ID', 'User_ID', 'user_nikname', 'user_xatm_number']
        sheet.append(headers)

        # Har bir sotuvchi uchun ma'lumotlarni Excel jadvalida yangi qatorga qo'shish
        for user in user_info:
            sheet.append(user)  # `user` ro'yxat bo'lishi kerak, shuning uchun to'g'ridan-to'g'ri qo'shamiz

        # Excel faylini saqlash
        file_path = 'users.xlsx'
        try:
            workbook.save(file_path)
        except Exception as e:
            await msg.answer(f"Faylni saqlashda xatolik yuz berdi: {e}")
            return

        # Faylni yopish
        workbook.close()

        # Correctly using InputFile with the file path
        document = FSInputFile(file_path)
        
        # Sending the document
        try:
            await msg.bot.send_document(chat_id=msg.from_user.id, document=document)
        except Exception as e:
            await msg.answer(f"Faylni yuborishda xatolik yuz berdi: {e}")
    else:
        await msg.answer("Foydalanuvchilar topilmadi.")

    # Clean up by removing the file after sending
    if os.path.exists(file_path):
        os.remove(file_path)

